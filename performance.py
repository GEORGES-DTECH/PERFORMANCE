from tkinter import *

from tkinter import messagebox

import openpyxl as xl
import os

root = Tk()
root.title("performance")
root.configure(background="#005b96", width=650, height=600)
root.iconbitmap("C:/Users/user/Desktop/performance/performance.ico")
root.geometry("590x430")


def process_performance(event=None):
    try:
        workbook = xl.load_workbook("performance.xlsx")
        sheet = workbook['PERFORMANCE']
        for row in range(10, 65 + 1):
            cell = sheet.cell(row, 4)
            if cell.value is None:
                cell.value = 0
            interest = round(cell.value * 0.015)
            interest_cell = sheet.cell(row, 7)
            interest_cell.value = interest

            # shares_bf and shares cf is members total savings
            # total_repaid is the members payment either on loan or savings at a particular month
            # loan_bf/cf is the given loan balance as at a specific datec cf(cf/bf=carried/brought foward)
            # fines =extra charges,penalties.
            # principle= required loan repayment
            # months_share=savings at a particular month or day
            # advance=short term loan repayable in one month at an interest of 10%
            # loan_category=available loans and their repayment plans
            # interest=charged on loan given at a rate of 1.5%
            # total repaid should always meet the required payments
            # dividend income on earnings ie fund
            shares_bf_cell = sheet.cell(row, 3)

            loan_bf_cell = sheet.cell(row, 4)

            total_repaid_cell = sheet.cell(row, 5)

            principle_cell = sheet.cell(row, 6)

            months_share_cell = sheet.cell(row, 8)

            shares_cf_cell = sheet.cell(row, 10)

            loan_cf_cell = sheet.cell(row, 11)

            fines_cell = sheet.cell(row, 9)

            adjustment_cell = sheet.cell(row, 2)

            advance_cell = sheet.cell(row, 12)

            loan_category_cell = sheet.cell(row, 13)

            dividend_cell = sheet.cell(row, 14)

            withdrawal_cell = sheet.cell(row, 18)
            if total_repaid_cell.value is None:
                total_repaid_cell.value = 0

            if loan_bf_cell.value is None:
                loan_bf_cell.value = 0

            if advance_cell.value is None:
                advance_cell.value = 0

            if shares_bf_cell.value is None:
                shares_bf_cell.value = 0

            if dividend_cell.value is None:
                dividend_cell.value = 0

            if dividend_cell.value > 1:
                total_repaid_cell.value = total_repaid_cell.value + dividend_cell.value

            if loan_category_cell.value is None:
                loan_category_cell.value = 0

            if withdrawal_cell.value is None:
                withdrawal_cell.value = 0

            if principle_cell.value is None:
                principle_cell.value = 0

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 5000:
                principle_cell.value = 350

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 10000:
                principle_cell.value = 500

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 20000:
                principle_cell.value = 1000

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 30000:
                principle_cell.value = 1500

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 50000:
                principle_cell.value = 2000

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 70000:
                principle_cell.value = 2800

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 100000:
                principle_cell.value = 3000

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 150000:
                principle_cell.value = 4500

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 200000:
                principle_cell.value = 6000

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 300000:
                principle_cell.value = 9000

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 500000:
                principle_cell.value = 15000

            if loan_bf_cell.value > 1 and shares_bf_cell.value > 1 and principle_cell.value == 0 and loan_category_cell.value == 1000000:
                principle_cell.value = 25000

            if fines_cell.value is None:
                fines_cell.value = 0
            if months_share_cell.value is None:
                months_share_cell.value = 0
            if shares_cf_cell.value is None:
                shares_cf_cell.value = 0
            if loan_cf_cell.value is None:
                loan_cf_cell.value = 0

            if adjustment_cell.value is None:
                adjustment_cell.value = 0

            if total_repaid_cell.value <= principle_cell.value:
                months_share_cell.value = total_repaid_cell.value
                share_cf = (total_repaid_cell.value + shares_bf_cell.value) - (
                        principle_cell.value + interest_cell.value)

                shares_cf_cell.value = share_cf
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value
                adjustments = principle_cell.value + interest_cell.value
                adjustment_cell.value = adjustments

            if total_repaid_cell.value <= principle_cell.value and fines_cell.value > 0:
                months_share_cell.value = total_repaid_cell.value
                share_cf = (total_repaid_cell.value + shares_bf_cell.value) - (
                        principle_cell.value + interest_cell.value)
                shares_cf_cell.value = share_cf
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value
                adjustments = principle_cell.value + interest_cell.value

                adjustment_cell.value = adjustments

            if total_repaid_cell.value <= principle_cell.value and fines_cell.value < 0:
                months_share_cell.value = total_repaid_cell.value
                share_cf = (total_repaid_cell.value + shares_bf_cell.value) - (
                        principle_cell.value + interest_cell.value + abs(fines_cell.value))

                shares_cf_cell.value = share_cf
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value
                adjustments = principle_cell.value + interest_cell.value + abs(fines_cell.value)

                adjustment_cell.value = adjustments

            if total_repaid_cell.value <= principle_cell.value and advance_cell.value > 1 and fines_cell.value < 0:
                months_share_cell.value = total_repaid_cell.value
                adjustment_cell.value = principle_cell.value + interest_cell.value + abs(
                    fines_cell.value) + advance_cell.value
                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.value - adjustment_cell.value
                adjustment_cell.value = adjustment_cell.value

            if total_repaid_cell.value <= principle_cell.value and advance_cell.value > 1 and fines_cell.value > 0:
                months_share_cell.value = total_repaid_cell.value
                adjustment_cell.value = principle_cell.value + interest_cell.value + advance_cell.value

                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.value - adjustment_cell.value
                adjustment_cell.value = adjustment_cell.value

            if total_repaid_cell.value <= principle_cell.value and advance_cell.value > 1 and fines_cell.value == 0:
                months_share_cell.value = total_repaid_cell.value
                adjustment_cell.value = principle_cell.value + interest_cell.value + advance_cell.value

                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.value - adjustment_cell.value

            if total_repaid_cell.value > principle_cell.value:
                months_share_cell.value = total_repaid_cell.value - principle_cell.value - interest_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value

            if total_repaid_cell.value > principle_cell.value and fines_cell.value < 0:
                months_share_cell.value = total_repaid_cell.value - principle_cell.value - interest_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value - abs(fines_cell.value)
                adjustment_cell.value = abs(fines_cell.value)

                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value

            if total_repaid_cell.value > principle_cell.value and fines_cell.value > 0:
                months_share_cell.value = total_repaid_cell.value - principle_cell.value - interest_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value

            if principle_cell.value < total_repaid_cell.value < advance_cell.value and fines_cell.value > 0:
                months_share_cell.value = total_repaid_cell.value - principle_cell.value - interest_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value - advance_cell.value
                adjustment_cell.value = advance_cell.value
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value

            if principle_cell.value < total_repaid_cell.value < advance_cell.value and fines_cell.value < 0:
                months_share_cell.value = total_repaid_cell.value - principle_cell.value - interest_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value - advance_cell.value - abs(
                    fines_cell.value)
                adjustment_cell.value = advance_cell.value + abs(fines_cell.value)

                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value

            if total_repaid_cell.value > advance_cell.value and total_repaid_cell.value - advance_cell.value > principle_cell.value and fines_cell.value < 0:
                total_repaid_cell.value = total_repaid_cell.value - advance_cell.value
                months_share_cell.value = total_repaid_cell.value - principle_cell.value - interest_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value - abs(fines_cell.value)
                adjustment_cell.value = abs(fines_cell.value)

            if total_repaid_cell.value > advance_cell.value and total_repaid_cell.value - advance_cell.value > principle_cell.value and fines_cell.value > 0:
                total_repaid_cell.value = total_repaid_cell.value - advance_cell.value
                months_share_cell.value = total_repaid_cell.value - principle_cell.value - interest_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value

            if principle_cell.value < total_repaid_cell.value < advance_cell.value and fines_cell.value == 0:
                months_share_cell.value = total_repaid_cell.value - principle_cell.value - interest_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value - advance_cell.value
                adjustment_cell.value = advance_cell.value
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value

            if total_repaid_cell.value > advance_cell.value and total_repaid_cell.value > principle_cell.value and fines_cell.value == 0:
                total_repaid_cell.value = total_repaid_cell.value - advance_cell.value
                months_share_cell.value = total_repaid_cell.value - principle_cell.value - interest_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value > 1:
                months_share_cell.value = total_repaid_cell.value
                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.value

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and fines_cell.value > 0 and total_repaid_cell.value > 1:
                months_share_cell.value = total_repaid_cell.value
                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.value

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and fines_cell.value < 0 and total_repaid_cell.value > 1:
                months_share_cell.value = total_repaid_cell.value
                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.value - abs(fines_cell.value)
                adjustment_cell.value = abs(fines_cell.value)

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value > 1 and \
                    1 < advance_cell.value <= total_repaid_cell.value:
                months_share_cell.value = total_repaid_cell.value
                months_share_cell.value = months_share_cell.value - advance_cell.value
                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.ife

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value > 1 and \
                    1 < advance_cell.value <= total_repaid_cell.value and fines_cell.value > 0:
                months_share_cell.value = total_repaid_cell.value
                months_share_cell.value = months_share_cell.value - advance_cell.value
                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.value

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value > 1 and \
                    1 < advance_cell.value <= total_repaid_cell.value and fines_cell.value < 0:
                months_share_cell.value = total_repaid_cell.value
                months_share_cell.value = months_share_cell.value - advance_cell.value
                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.value - abs(fines_cell.value)
                adjustment_cell.value = abs(fines_cell.value)

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value > 1 and \
                    1 < advance_cell.value <= total_repaid_cell.value and fines_cell.value == 0:
                months_share_cell.value = total_repaid_cell.value
                months_share_cell.value = months_share_cell.value - advance_cell.value
                shares_cf_cell.value = months_share_cell.value + shares_bf_cell.value

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value > 1 and \
                    advance_cell.value > 1 and total_repaid_cell.value <= advance_cell.value and fines_cell.value > 0:
                months_share_cell.value = total_repaid_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value - advance_cell.value
                adjustment_cell.value = advance_cell.value

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value > 1 and \
                    advance_cell.value > 1 and total_repaid_cell.value <= advance_cell.value and fines_cell.value < 0 \
                    :
                months_share_cell.value = total_repaid_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value - advance_cell.value - abs(
                    fines_cell.value)
                adjustment_cell.value = advance_cell.value + abs(fines_cell.value)

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value > 1 and \
                    advance_cell.value > 1 and total_repaid_cell.value <= advance_cell.value and fines_cell.value == 0 \
                    :
                months_share_cell.value = total_repaid_cell.value
                shares_cf_cell.value = shares_bf_cell.value + months_share_cell.value - advance_cell.value

                adjustment_cell.value = advance_cell.value

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value == 0:
                shares_cf_cell.value = shares_bf_cell.value - abs(fines_cell.value)
                adjustment_cell.value = abs(fines_cell.value)

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value == 0 and \
                    advance_cell.value > 1 and fines_cell.value == 0:
                months_share_cell.value = total_repaid_cell.value
                shares_cf_cell.value = shares_bf_cell.value - advance_cell.value
                adjustment_cell.value = advance_cell.value

            if shares_bf_cell.value > 1 and loan_bf_cell.value == 0 and total_repaid_cell.value == 0 and \
                    advance_cell.value > 1 and fines_cell.value < 0:
                months_share_cell.value = total_repaid_cell.value
                shares_cf_cell.value = shares_bf_cell.value - advance_cell.value - abs(fines_cell.value)
                adjustment_cell.value = advance_cell.value + abs(fines_cell.value)

            if shares_bf_cell.value == 0 and loan_bf_cell.value >= 1 and total_repaid_cell.value > 1 and advance_cell.value == 0 and \
                    fines_cell.value < 0:
                principle_cell.value = total_repaid_cell.value - interest_cell.value
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value - abs(fines_cell.value)
                shares_cf_cell.value = shares_bf_cell.value
                months_share_cell.value = 0
                adjustment_cell.value = "DEFAULTER"

            if shares_bf_cell.value == 0 and loan_bf_cell.value >= 1 and total_repaid_cell.value == 0 and advance_cell.value == 0:
                principle_cell.value = abs(fines_cell.value) - interest_cell.value
                loan_cf_cell.value = loan_bf_cell.value - principle_cell.value
                months_share_cell.value = 0
                shares_cf_cell.value = 0
                adjustment_cell.value = "DEFAULTER"
            if withdrawal_cell.value > 1:
                shares_cf_cell.value = shares_cf_cell.value - withdrawal_cell.value
            # summary formulas
            sheet["b2"] = '=sum(j10:j65)'
            sheet["b3"] = '=sum(L10:L65)'
            sheet["b4"] = '=sum(j71:j89)+sum(k10:k66)'
            sheet["b6"] = '=sum(b3:b5)'
            sheet["b67"] = '=sum(b10:b66)'
            sheet["c67"] = '=sum(c10:c66)'
            sheet["d67"] = '=sum(d10:d66)'
            sheet["e67"] = '=sum(e10:e66)'
            sheet["f67"] = '=sum(f10:f66)'
            sheet["g67"] = '=sum(g10:g66)'
            sheet["h67"] = '=sum(h10:h66)'
            sheet["i67"] = '=sum(i10:i66)'
            sheet["j67"] = '=sum(j10:j66)'
            sheet["k67"] = '=sum(k10:k66)'
            sheet["l67"] = '=sum(l10:l66)'
            sheet["N67"] = '=sum(N10:N66)'
            sheet["R67"] = '=sum(R10:R66)'
            sheet["c70"] = '=sum(e10:e66)-sum(n10:n65)'
            sheet["c71"] = '=sum(l10:l66)'
            sheet["c79"] = '=sum(c70:c78)'
            sheet["e70"] = '=sum(h71:h89)'
            sheet["e71"] = '=sum(k71:k89)'
            sheet["e72"] = '=sum(j71:j89)'
            sheet["e79"] = '=sum(e70:e78)'

            workbook.save("performance.xlsx")
        response = messagebox.showinfo(title="performance processing", message="finished")
        if response == "Ok":
            response.quit()

    except PermissionError:
        messagebox.showerror(title="error", message="close the performance and try again")


PERFORMANCE_FRAME = Frame(root, background="green", width=300, height=300).grid(row=5, column=1)
REMINDER = Label(PERFORMANCE_FRAME, text=" button changes from red to white when finished", background="green",
                 foreground="black", font='Helvetica 10 bold')
REMINDER.grid(row=5, column=1)

PERFORMANCE_BUTTON = Button(root, text="process performance", bd=5, activebackground="red", command=process_performance)
PERFORMANCE_BUTTON.grid(row=0, column=2)
root.bind("<Tab>", process_performance)


def bring_forward():
    try:
        workbook = xl.load_workbook('performance.xlsx')
        sheet = workbook['PERFORMANCE']
        for row in range(10, 65 + 1):
            interest_cell = sheet.cell(row, 7)
            shares_bf_cell = sheet.cell(row, 3)

            loan_bf_cell = sheet.cell(row, 4)
            total_repaid_cell = sheet.cell(row, 5)

            principle_cell = sheet.cell(row, 6)

            months_share_cell = sheet.cell(row, 8)

            shares_cf_cell = sheet.cell(row, 10)

            loan_cf_cell = sheet.cell(row, 11)

            fines_cell = sheet.cell(row, 9)

            adjustment_cell = sheet.cell(row, 2)

            advance_cell = sheet.cell(row, 12)

            loan_category_cell = sheet.cell(row, 13)

            withdrawal_cell = sheet.cell(row, 18)
            if withdrawal_cell.value is None:
                withdrawal_cell.value = 0
            if loan_category_cell.value is None:
                loan_category_cell.value = 0

            if total_repaid_cell.value is None:
                total_repaid_cell.value = 0
            if shares_bf_cell.value is None:
                shares_bf_cell.value = 0
            if principle_cell.value is None:
                principle_cell.value = 0
            if fines_cell.value is None:
                fines_cell.value = 0
            if months_share_cell.value is None:
                months_share_cell.value = 0
            if shares_cf_cell.value is None:
                shares_cf_cell.value = 0
            if loan_cf_cell.value is None:
                loan_cf_cell.value = 0
            if loan_bf_cell.value is None:
                loan_bf_cell.value = 0
            if adjustment_cell.value is None:
                adjustment_cell.value = 0
            if advance_cell.value is None:
                advance_cell.value = 0
            if interest_cell.value is None:
                interest_cell.value = 0
            shares_bf_cell.value = shares_cf_cell.value
            loan_bf_cell.value = loan_cf_cell.value
            interest_cell.value = round(loan_bf_cell.value * 0.015)
            loan_bf_cell.value = loan_cf_cell.value
            advance_cell.value = (advance_cell.value * 0.1) + advance_cell.value
            total_repaid_cell.value = 0
            principle_cell.value = 0
            months_share_cell.value = 0
            shares_cf_cell.value = 0
            loan_cf_cell.value = 0
            adjustment_cell.value = 0
            fines_cell.value = 0
            withdrawal_cell.value = 0

        workbook.save('performance.xlsx')

        response = messagebox.showinfo(title="forward", message="finished ")
        if response == "Ok":
            response.quit()
    except PermissionError:

        messagebox.showerror(title="error", message="close the performance and try again")


FOWARD_BUTTON = Button(root, text="bring foward", bd=5, activebackground="red", command=bring_forward)
FOWARD_BUTTON.grid(row=5, column=2)


def quit():
    iexit = messagebox.askyesno(title="quit", message="DO YOU WISH TO QUIT")
    if iexit > 0:
        root.destroy()
        return


QUIT_BUTTON = Button(root, text="Quit program", bd=5, command=quit)
QUIT_BUTTON.grid(row=6, column=2)


def open_performance():
   os.startfile("performance.xlsx")




OPENBUTTON = Button(root, text="open performance", bd=5, command=open_performance, activebackground="red")
OPENBUTTON.grid(row=5, column=0)


def dividendreturn(event):
    DIVIDEND.focus()


def averagereturn(event):
    AVERAGE.focus()


Dividend = IntVar()
Average = IntVar()
AVERAGE = Entry(root, bd=5, font="helvetica 14 bold", textvariable=Average)
AVERAGE.grid(row=6, column=1)
AVERAGE.bind("<Up>", dividendreturn)

DIVIDEND = Entry(root, font="helvetica 14 bold", bd=5, textvariable=Dividend)
DIVIDEND.grid(row=4, column=1)
DIVIDEND.bind("<Down>", averagereturn)

DIVIDENDLABLE = Label(root, text="DIVIDEND AMOUNT", bd=5).grid(row=4, column=0)
AVERAGELABLE = Label(root, text="TOTAL AVERAGE", bd=5).grid(row=6, column=0)


def dividend_calculation(event=None):
    try:
        workbook = xl.load_workbook("performance.xlsx")
        sheet = workbook["DIVIDEND CALCULATION"]

        for row in range(5, 64 + 1):

            totalsyear1_cell = sheet.cell(row, 14)
            average_cell = sheet.cell(row, 15)
            month1_cell = sheet.cell(row, 2)
            month2_cell = sheet.cell(row, 3)
            month3_cell = sheet.cell(row, 4)
            month4_cell = sheet.cell(row, 5)
            month5_cell = sheet.cell(row, 6)
            month6_cell = sheet.cell(row, 7)
            month7_cell = sheet.cell(row, 8)
            month8_cell = sheet.cell(row, 9)
            month9_cell = sheet.cell(row, 10)
            month10_cell = sheet.cell(row, 11)
            month11_cell = sheet.cell(row, 12)
            month12_cell = sheet.cell(row, 13)

            if month1_cell.value is None:
                month1_cell.value = 0
            if month2_cell.value is None:
                month2_cell.value = 0
            if month3_cell.value is None:
                month3_cell.value = 0
            if month4_cell.value is None:
                month4_cell.value = 0
            if month5_cell.value is None:
                month5_cell.value = 0
            if month6_cell.value is None:
                month6_cell.value = 0
            if month7_cell.value is None:
                month7_cell.value = 0

            if month8_cell.value is None:
                month8_cell.value = 0

            if month9_cell.value is None:
                month9_cell.value = 0
            if month10_cell.value is None:
                month10_cell.value = 0
            if month11_cell.value is None:
                month11_cell.value = 0
            if month12_cell.value is None:
                month12_cell.value = 0
            if totalsyear1_cell.value is None:
                totalsyear1_cell.value = 0
            if average_cell.value is None:
                average_cell.value = 0

            totalsyear1_cell.value = month1_cell.value + month2_cell.value + month3_cell.value + month4_cell.value + month5_cell.value + \
                                     month6_cell.value + month7_cell.value + month8_cell.value + month9_cell.value + month10_cell.value + month11_cell.value + month12_cell.value
            average_cell.value = round(totalsyear1_cell.value / 12)

            totalsyear2_cell = sheet.cell(row, 31)
            averageyear2_cell = sheet.cell(row, 32)
            month1_cell = sheet.cell(row, 19)
            month2_cell = sheet.cell(row, 20)
            month3_cell = sheet.cell(row, 21)
            month4_cell = sheet.cell(row, 22)
            month5_cell = sheet.cell(row, 23)
            month6_cell = sheet.cell(row, 24)
            month7_cell = sheet.cell(row, 25)
            month8_cell = sheet.cell(row, 26)
            month9_cell = sheet.cell(row, 27)
            month10_cell = sheet.cell(row, 28)
            month11_cell = sheet.cell(row, 29)
            month12_cell = sheet.cell(row, 30)

            if month1_cell.value is None:
                month1_cell.value = 0
            if month2_cell.value is None:
                month2_cell.value = 0
            if month3_cell.value is None:
                month3_cell.value = 0
            if month4_cell.value is None:
                month4_cell.value = 0
            if month5_cell.value is None:
                month5_cell.value = 0
            if month6_cell.value is None:
                month6_cell.value = 0
            if month7_cell.value is None:
                month7_cell.value = 0

            if month8_cell.value is None:
                month8_cell.value = 0

            if month9_cell.value is None:
                month9_cell.value = 0
            if month10_cell.value is None:
                month10_cell.value = 0
            if month11_cell.value is None:
                month11_cell.value = 0
            if month12_cell.value is None:
                month12_cell.value = 0
            if totalsyear2_cell.value is None:
                totalsyear2_cell.value = 0
            if averageyear2_cell.value is None:
                averageyear2_cell.value = 0
            totalsyear2_cell.value = month1_cell.value + month2_cell.value + month3_cell.value + month4_cell.value + month5_cell.value + \
                                     month6_cell.value + month7_cell.value + month8_cell.value + month9_cell.value + month10_cell.value + month11_cell.value + month12_cell.value

            averageyear2_cell.value = round(totalsyear2_cell.value / 12)

            totalsyear3_cell = sheet.cell(row, 48)
            averageyear3_cell = sheet.cell(row, 49)
            if totalsyear3_cell.value is None:
                totalsyear3_cell.value = 0
            if averageyear3_cell.value is None:
                averageyear3_cell.value = 0

            month1_cell = sheet.cell(row, 36)
            month2_cell = sheet.cell(row, 37)
            month3_cell = sheet.cell(row, 38)
            month4_cell = sheet.cell(row, 39)
            month5_cell = sheet.cell(row, 40)
            month6_cell = sheet.cell(row, 41)
            month7_cell = sheet.cell(row, 42)
            month8_cell = sheet.cell(row, 43)
            month9_cell = sheet.cell(row, 44)
            month10_cell = sheet.cell(row, 45)
            month11_cell = sheet.cell(row, 46)
            month12_cell = sheet.cell(row, 47)

            if month1_cell.value is None:
                month1_cell.value = 0
            if month2_cell.value is None:
                month2_cell.value = 0
            if month3_cell.value is None:
                month3_cell.value = 0
            if month4_cell.value is None:
                month4_cell.value = 0
            if month5_cell.value is None:
                month5_cell.value = 0
            if month6_cell.value is None:
                month6_cell.value = 0
            if month7_cell.value is None:
                month7_cell.value = 0

            if month8_cell.value is None:
                month8_cell.value = 0

            if month9_cell.value is None:
                month9_cell.value = 0
            if month10_cell.value is None:
                month10_cell.value = 0
            if month11_cell.value is None:
                month11_cell.value = 0
            if month12_cell.value is None:
                month12_cell.value = 0
            totalsyear3_cell.value = month1_cell.value + month2_cell.value + month3_cell.value + month4_cell.value + month5_cell.value + \
                                     month6_cell.value + month7_cell.value + month8_cell.value + month9_cell.value + month10_cell.value + month11_cell.value + month12_cell.value

            averageyear3_cell.value = round(totalsyear3_cell.value / 12)

            totalsyear4_cell = sheet.cell(row, 65)
            averageyear4_cell = sheet.cell(row, 66)

            month1_cell = sheet.cell(row, 53)
            month2_cell = sheet.cell(row, 54)
            month3_cell = sheet.cell(row, 55)
            month4_cell = sheet.cell(row, 56)
            month5_cell = sheet.cell(row, 57)
            month6_cell = sheet.cell(row, 58)
            month7_cell = sheet.cell(row, 59)
            month8_cell = sheet.cell(row, 60)
            month9_cell = sheet.cell(row, 61)
            month10_cell = sheet.cell(row, 62)
            month11_cell = sheet.cell(row, 63)
            month12_cell = sheet.cell(row, 64)

            if month1_cell.value is None:
                month1_cell.value = 0
            if month2_cell.value is None:
                month2_cell.value = 0
            if month3_cell.value is None:
                month3_cell.value = 0
            if month4_cell.value is None:
                month4_cell.value = 0
            if month5_cell.value is None:
                month5_cell.value = 0
            if month6_cell.value is None:
                month6_cell.value = 0
            if month7_cell.value is None:
                month7_cell.value = 0

            if month8_cell.value is None:
                month8_cell.value = 0

            if month9_cell.value is None:
                month9_cell.value = 0
            if month10_cell.value is None:
                month10_cell.value = 0
            if month11_cell.value is None:
                month11_cell.value = 0
            if month12_cell.value is None:
                month12_cell.value = 0
            if totalsyear4_cell.value is None:
                totalsyear4_cell.value = 0
            if averageyear4_cell.value is None:
                averageyear4_cell.value = 0

            totalsyear4_cell.value = month1_cell.value + month2_cell.value + month3_cell.value + month4_cell.value + month5_cell.value + \
                                     month6_cell.value + month7_cell.value + month8_cell.value + month9_cell.value + month10_cell.value + month11_cell.value + month12_cell.value
            averageyear4_cell.value = round(totalsyear4_cell.value / 12)

            totalsyear5_cell = sheet.cell(row, 82)
            averageyear5_cell = sheet.cell(row, 83)

            month1_cell = sheet.cell(row, 70)
            month2_cell = sheet.cell(row, 71)
            month3_cell = sheet.cell(row, 72)
            month4_cell = sheet.cell(row, 73)
            month5_cell = sheet.cell(row, 74)
            month6_cell = sheet.cell(row, 75)
            month7_cell = sheet.cell(row, 76)
            month8_cell = sheet.cell(row, 77)
            month9_cell = sheet.cell(row, 78)
            month10_cell = sheet.cell(row, 79)
            month11_cell = sheet.cell(row, 80)
            month12_cell = sheet.cell(row, 81)

            if month1_cell.value is None:
                month1_cell.value = 0
            if month2_cell.value is None:
                month2_cell.value = 0
            if month3_cell.value is None:
                month3_cell.value = 0
            if month4_cell.value is None:
                month4_cell.value = 0
            if month5_cell.value is None:
                month5_cell.value = 0
            if month6_cell.value is None:
                month6_cell.value = 0
            if month7_cell.value is None:
                month7_cell.value = 0

            if month8_cell.value is None:
                month8_cell.value = 0

            if month9_cell.value is None:
                month9_cell.value = 0
            if month10_cell.value is None:
                month10_cell.value = 0
            if month11_cell.value is None:
                month11_cell.value = 0
            if month12_cell.value is None:
                month12_cell.value = 0
            if totalsyear5_cell.value is None:
                totalsyear5_cell.value = 0
            if averageyear5_cell.value is None:
                averageyear5_cell.value = 0

            totalsyear5_cell.value = month1_cell.value + month2_cell.value + month3_cell.value + month4_cell.value + month5_cell.value + \
                                     month6_cell.value + month7_cell.value + month8_cell.value + month9_cell.value + month10_cell.value + month11_cell.value + month12_cell.value
            averageyear5_cell.value = round(totalsyear5_cell.value / 12)

            totalsyear6_cell = sheet.cell(row, 99)
            averageyear6_cell = sheet.cell(row, 100)

            month1_cell = sheet.cell(row, 87)
            month2_cell = sheet.cell(row, 88)
            month3_cell = sheet.cell(row, 89)
            month4_cell = sheet.cell(row, 90)
            month5_cell = sheet.cell(row, 91)
            month6_cell = sheet.cell(row, 92)
            month7_cell = sheet.cell(row, 93)
            month8_cell = sheet.cell(row, 94)
            month9_cell = sheet.cell(row, 95)
            month10_cell = sheet.cell(row, 96)
            month11_cell = sheet.cell(row, 97)
            month12_cell = sheet.cell(row, 98)

            if month1_cell.value is None:
                month1_cell.value = 0
            if month2_cell.value is None:
                month2_cell.value = 0
            if month3_cell.value is None:
                month3_cell.value = 0
            if month4_cell.value is None:
                month4_cell.value = 0
            if month5_cell.value is None:
                month5_cell.value = 0
            if month6_cell.value is None:
                month6_cell.value = 0
            if month7_cell.value is None:
                month7_cell.value = 0

            if month8_cell.value is None:
                month8_cell.value = 0

            if month9_cell.value is None:
                month9_cell.value = 0
            if month10_cell.value is None:
                month10_cell.value = 0
            if month11_cell.value is None:
                month11_cell.value = 0
            if month12_cell.value is None:
                month12_cell.value = 0
            if totalsyear6_cell.value is None:
                totalsyear6_cell.value = 0
            if averageyear6_cell.value is None:
                averageyear6_cell.value = 0

            totalsyear6_cell.value = month1_cell.value + month2_cell.value + month3_cell.value + month4_cell.value + month5_cell.value + \
                                     month6_cell.value + month7_cell.value + month8_cell.value + month9_cell.value + month10_cell.value + month11_cell.value + month12_cell.value

            averageyear6_cell.value = round(totalsyear6_cell.value / 12)

            totalsyear7_cell = sheet.cell(row, 116)
            averageyear7_cell = sheet.cell(row, 117)

            month1_cell = sheet.cell(row, 104)
            month2_cell = sheet.cell(row, 105)
            month3_cell = sheet.cell(row, 106)
            month4_cell = sheet.cell(row, 107)
            month5_cell = sheet.cell(row, 108)
            month6_cell = sheet.cell(row, 109)
            month7_cell = sheet.cell(row, 110)
            month8_cell = sheet.cell(row, 111)
            month9_cell = sheet.cell(row, 112)
            month10_cell = sheet.cell(row, 113)
            month11_cell = sheet.cell(row, 114)
            month12_cell = sheet.cell(row, 115)

            if month1_cell.value is None:
                month1_cell.value = 0
            if month2_cell.value is None:
                month2_cell.value = 0
            if month3_cell.value is None:
                month3_cell.value = 0
            if month4_cell.value is None:
                month4_cell.value = 0
            if month5_cell.value is None:
                month5_cell.value = 0
            if month6_cell.value is None:
                month6_cell.value = 0
            if month7_cell.value is None:
                month7_cell.value = 0

            if month8_cell.value is None:
                month8_cell.value = 0

            if month9_cell.value is None:
                month9_cell.value = 0
            if month10_cell.value is None:
                month10_cell.value = 0
            if month11_cell.value is None:
                month11_cell.value = 0
            if month12_cell.value is None:
                month12_cell.value = 0
            if totalsyear7_cell.value is None:
                totalsyear7_cell.value = 0
            if averageyear7_cell.value is None:
                averageyear7_cell.value = 0

            totalsyear7_cell.value = month1_cell.value + month2_cell.value + month3_cell.value + month4_cell.value + month5_cell.value + \
                                     month6_cell.value + month7_cell.value + month8_cell.value + month9_cell.value + month10_cell.value + month11_cell.value + month12_cell.value
            averageyear7_cell.value = round(totalsyear7_cell.value / 12)

            total_average_cell = sheet.cell(row, 138)
            dividend_cell = sheet.cell(row, 139)
            totalsyear8_cell = sheet.cell(row, 133)
            averageyear8_cell = sheet.cell(row, 134)

            month1_cell = sheet.cell(row, 121)
            month2_cell = sheet.cell(row, 122)
            month3_cell = sheet.cell(row, 123)
            month4_cell = sheet.cell(row, 124)
            month5_cell = sheet.cell(row, 125)
            month6_cell = sheet.cell(row, 126)
            month7_cell = sheet.cell(row, 127)
            month8_cell = sheet.cell(row, 128)
            month9_cell = sheet.cell(row, 129)
            month10_cell = sheet.cell(row, 130)
            month11_cell = sheet.cell(row, 131)
            month12_cell = sheet.cell(row, 132)

            if month1_cell.value is None:
                month1_cell.value = 0
            if month2_cell.value is None:
                month2_cell.value = 0
            if month3_cell.value is None:
                month3_cell.value = 0
            if month4_cell.value is None:
                month4_cell.value = 0
            if month5_cell.value is None:
                month5_cell.value = 0
            if month6_cell.value is None:
                month6_cell.value = 0
            if month7_cell.value is None:
                month7_cell.value = 0

            if month8_cell.value is None:
                month8_cell.value = 0

            if month9_cell.value is None:
                month9_cell.value = 0
            if month10_cell.value is None:
                month10_cell.value = 0
            if month11_cell.value is None:
                month11_cell.value = 0
            if month12_cell.value is None:
                month12_cell.value = 0
            if totalsyear8_cell.value is None:
                totalsyear8_cell.value = 0
            if averageyear8_cell.value is None:
                averageyear8_cell.value = 0
            if dividend_cell.value is None:
                dividend_cell.value = 0
            if total_average_cell.value is None:
                total_average_cell.value = 0

            totalsyear8_cell.value = month1_cell.value + month2_cell.value + month3_cell.value + month4_cell.value + month5_cell.value + \
                                     month6_cell.value + month7_cell.value + month8_cell.value + month9_cell.value + month10_cell.value + month11_cell.value + month12_cell.value
            averageyear8_cell.value = round(totalsyear8_cell.value / 12)
            total_average_cell.value = average_cell.value + averageyear2_cell.value + averageyear3_cell.value + averageyear4_cell.value + \
                                       averageyear5_cell.value + averageyear6_cell.value + averageyear7_cell.value + averageyear8_cell.value

            if Dividend.get() > 1 and Average.get() > 1:
                dividend_cell.value = round(Dividend.get() / Average.get() * total_average_cell.value)

        workbook.save("performance.xlsx")
        response = messagebox.showinfo(title="finished", message="finished ")
        if response == "Ok":
            response.quit()
    except PermissionError:
        messagebox.showerror(title="error", message="close the performance and try again")

    except ValueError:
        messagebox.showerror(title="error", message="input number 0 in divivend and average entry area")


DIVIDENDBUTTON = Button(root, text="calculate dividend", bd=3, activebackground="red", command=dividend_calculation)
DIVIDENDBUTTON.grid(row=0, column=1)
root.bind("<Return>", dividend_calculation)


def reset_dividendsheet(event=None):
    try:
        workbook = xl.load_workbook("performance.xlsx")
        sheet = workbook["DIVIDEND CALCULATION"]

        for row in range(5, 64 + 1):
            total_average_cell = sheet.cell(row, 138)
            dividend_cell = sheet.cell(row, 139)
            totalsyear1_cell = sheet.cell(row, 14)
            averageyear1_cell = sheet.cell(row, 15)
            month1_cell = sheet.cell(row, 2)
            month2_cell = sheet.cell(row, 3)
            month3_cell = sheet.cell(row, 4)
            month4_cell = sheet.cell(row, 5)
            month5_cell = sheet.cell(row, 6)
            month6_cell = sheet.cell(row, 7)
            month7_cell = sheet.cell(row, 8)
            month8_cell = sheet.cell(row, 9)
            month9_cell = sheet.cell(row, 10)
            month10_cell = sheet.cell(row, 11)
            month11_cell = sheet.cell(row, 12)
            month12_cell = sheet.cell(row, 13)

            month1_cell.value = 0
            month2_cell.value = 0
            month3_cell.value = 0
            month4_cell.value = 0
            month5_cell.value = 0
            month6_cell.value = 0
            month7_cell.value = 0
            month8_cell.value = 0
            month9_cell.value = 0
            month10_cell.value = 0
            month11_cell.value = 0
            month12_cell.value = 0
            totalsyear1_cell.value = 0
            averageyear1_cell.value = 0
            dividend_cell.value = 0
            total_average_cell.value = 0

            totalsyear2_cell = sheet.cell(row, 31)
            averageyear2_cell = sheet.cell(row, 32)
            month1_cell = sheet.cell(row, 19)
            month2_cell = sheet.cell(row, 20)
            month3_cell = sheet.cell(row, 21)
            month4_cell = sheet.cell(row, 22)
            month5_cell = sheet.cell(row, 23)
            month6_cell = sheet.cell(row, 24)
            month7_cell = sheet.cell(row, 25)
            month8_cell = sheet.cell(row, 26)
            month9_cell = sheet.cell(row, 27)
            month10_cell = sheet.cell(row, 28)
            month11_cell = sheet.cell(row, 29)
            month12_cell = sheet.cell(row, 30)
            month1_cell.value = 0
            month2_cell.value = 0
            month3_cell.value = 0
            month4_cell.value = 0
            month5_cell.value = 0
            month6_cell.value = 0
            month7_cell.value = 0
            month8_cell.value = 0
            month9_cell.value = 0
            month10_cell.value = 0
            month11_cell.value = 0
            month12_cell.value = 0

            totalsyear2_cell.value = 0
            averageyear2_cell.value = 0

            totalsyear3_cell = sheet.cell(row, 48)
            averageyear3_cell = sheet.cell(row, 49)
            month1_cell = sheet.cell(row, 36)
            month2_cell = sheet.cell(row, 37)
            month3_cell = sheet.cell(row, 38)
            month4_cell = sheet.cell(row, 39)
            month5_cell = sheet.cell(row, 40)
            month6_cell = sheet.cell(row, 41)
            month7_cell = sheet.cell(row, 42)
            month8_cell = sheet.cell(row, 43)
            month9_cell = sheet.cell(row, 44)
            month10_cell = sheet.cell(row, 45)
            month11_cell = sheet.cell(row, 46)
            month12_cell = sheet.cell(row, 47)
            month1_cell.value = 0
            month2_cell.value = 0
            month3_cell.value = 0
            month4_cell.value = 0
            month5_cell.value = 0
            month6_cell.value = 0
            month7_cell.value = 0
            month8_cell.value = 0
            month9_cell.value = 0
            month10_cell.value = 0
            month11_cell.value = 0
            month12_cell.value = 0

            totalsyear3_cell.value = 0
            averageyear3_cell.value = 0

            totalsyear4_cell = sheet.cell(row, 65)
            averageyear4_cell = sheet.cell(row, 66)

            month1_cell = sheet.cell(row, 53)
            month2_cell = sheet.cell(row, 54)
            month3_cell = sheet.cell(row, 55)
            month4_cell = sheet.cell(row, 56)
            month5_cell = sheet.cell(row, 57)
            month6_cell = sheet.cell(row, 58)
            month7_cell = sheet.cell(row, 59)
            month8_cell = sheet.cell(row, 60)
            month9_cell = sheet.cell(row, 61)
            month10_cell = sheet.cell(row, 62)
            month11_cell = sheet.cell(row, 63)
            month12_cell = sheet.cell(row, 64)
            month1_cell.value = 0
            month2_cell.value = 0
            month3_cell.value = 0
            month4_cell.value = 0
            month5_cell.value = 0
            month6_cell.value = 0
            month7_cell.value = 0
            month8_cell.value = 0
            month9_cell.value = 0
            month10_cell.value = 0
            month11_cell.value = 0
            month12_cell.value = 0

            totalsyear4_cell.value = 0
            averageyear4_cell.value = 0

            totalsyear5_cell = sheet.cell(row, 82)
            averageyear5_cell = sheet.cell(row, 83)

            month1_cell = sheet.cell(row, 70)
            month2_cell = sheet.cell(row, 71)
            month3_cell = sheet.cell(row, 72)
            month4_cell = sheet.cell(row, 73)
            month5_cell = sheet.cell(row, 74)
            month6_cell = sheet.cell(row, 75)
            month7_cell = sheet.cell(row, 76)
            month8_cell = sheet.cell(row, 77)
            month9_cell = sheet.cell(row, 78)
            month10_cell = sheet.cell(row, 79)
            month11_cell = sheet.cell(row, 80)
            month12_cell = sheet.cell(row, 81)
            month1_cell.value = 0
            month2_cell.value = 0
            month3_cell.value = 0
            month4_cell.value = 0
            month5_cell.value = 0
            month6_cell.value = 0
            month7_cell.value = 0
            month8_cell.value = 0
            month9_cell.value = 0
            month10_cell.value = 0
            month11_cell.value = 0
            month12_cell.value = 0

            totalsyear5_cell.value = 0
            averageyear5_cell.value = 0

            totalsyear6_cell = sheet.cell(row, 99)
            averageyear6_cell = sheet.cell(row, 100)

            month1_cell = sheet.cell(row, 87)
            month2_cell = sheet.cell(row, 88)
            month3_cell = sheet.cell(row, 89)
            month4_cell = sheet.cell(row, 90)
            month5_cell = sheet.cell(row, 91)
            month6_cell = sheet.cell(row, 92)
            month7_cell = sheet.cell(row, 93)
            month8_cell = sheet.cell(row, 94)
            month9_cell = sheet.cell(row, 95)
            month10_cell = sheet.cell(row, 96)
            month11_cell = sheet.cell(row, 97)
            month12_cell = sheet.cell(row, 98)
            month1_cell.value = 0
            month2_cell.value = 0
            month3_cell.value = 0
            month4_cell.value = 0
            month5_cell.value = 0
            month6_cell.value = 0
            month7_cell.value = 0
            month8_cell.value = 0
            month9_cell.value = 0
            month10_cell.value = 0
            month11_cell.value = 0
            month12_cell.value = 0

            totalsyear6_cell.value = 0
            averageyear6_cell.value = 0

            totalsyear7_cell = sheet.cell(row, 116)
            averageyear7_cell = sheet.cell(row, 117)

            month1_cell = sheet.cell(row, 104)
            month2_cell = sheet.cell(row, 105)
            month3_cell = sheet.cell(row, 106)
            month4_cell = sheet.cell(row, 107)
            month5_cell = sheet.cell(row, 108)
            month6_cell = sheet.cell(row, 109)
            month7_cell = sheet.cell(row, 110)
            month8_cell = sheet.cell(row, 111)
            month9_cell = sheet.cell(row, 112)
            month10_cell = sheet.cell(row, 113)
            month11_cell = sheet.cell(row, 114)
            month12_cell = sheet.cell(row, 115)
            month1_cell.value = 0
            month2_cell.value = 0
            month3_cell.value = 0
            month4_cell.value = 0
            month5_cell.value = 0
            month6_cell.value = 0
            month7_cell.value = 0
            month8_cell.value = 0
            month9_cell.value = 0
            month10_cell.value = 0
            month11_cell.value = 0
            month12_cell.value = 0

            totalsyear7_cell.value = 0
            averageyear7_cell.value = 0

            totalsyear8_cell = sheet.cell(row, 133)
            averageyear8_cell = sheet.cell(row, 134)

            month1_cell = sheet.cell(row, 121)
            month2_cell = sheet.cell(row, 122)
            month3_cell = sheet.cell(row, 123)
            month4_cell = sheet.cell(row, 124)
            month5_cell = sheet.cell(row, 125)
            month6_cell = sheet.cell(row, 126)
            month7_cell = sheet.cell(row, 127)
            month8_cell = sheet.cell(row, 128)
            month9_cell = sheet.cell(row, 129)
            month10_cell = sheet.cell(row, 130)
            month11_cell = sheet.cell(row, 131)
            month12_cell = sheet.cell(row, 132)
            month1_cell.value = 0
            month2_cell.value = 0
            month3_cell.value = 0
            month4_cell.value = 0
            month5_cell.value = 0
            month6_cell.value = 0
            month7_cell.value = 0
            month8_cell.value = 0
            month9_cell.value = 0
            month10_cell.value = 0
            month11_cell.value = 0
            month12_cell.value = 0

            totalsyear8_cell.value = 0
            averageyear8_cell.value = 0

        workbook.save("performance.xlsx")
        response = messagebox.showinfo(title="finished", message="finished ")
        if response == "Ok":
            response.quit()
    except PermissionError:
        messagebox.showerror(title="error", message="close the performance and try again")


RESETBUTTON = Button(root, text="Reset dividend sheet", activebackground="red", bd=5, command=reset_dividendsheet)
RESETBUTTON.grid(row=0, column=0)
root.bind("<Delete>", reset_dividendsheet)
root.mainloop()
